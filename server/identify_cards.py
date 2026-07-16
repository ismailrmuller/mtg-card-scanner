#!/usr/bin/env python3
"""
MTG Card Identifier
Reads images from scans/, identifies both cards per image using Claude AI,
saves to SQLite + exports to Excel with thumbnail previews.

Usage:
    $env:ANTHROPIC_API_KEY = "sk-ant-..."
    python identify_cards.py

Options (edit the CONFIG block below):
    MODEL      : claude-haiku-4-5-20251001  (cheap, fast)
                 claude-sonnet-4-6           (more accurate on hard images)
    SCANS_DIR  : folder with .jpg images from the scanner
    OUTPUT_DIR : where to write cards.db, cards.xlsx, card_thumbs/
"""

import anthropic
import base64
import io
import json
import os
import sqlite3
import sys
import time
from io import BytesIO
from pathlib import Path

from PIL import Image as PilImage
import openpyxl
from rich.console import Console
from rich.progress import (BarColumn, MofNCompleteColumn, Progress,
                           SpinnerColumn, TextColumn, TimeElapsedColumn,
                           TimeRemainingColumn)

console = Console()
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── .env loader (no external dependency) ─────────────────────────────────────

def load_dotenv():
    env_file = Path(__file__).parent / ".env"
    if not env_file.is_file():
        return
    for line in env_file.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, _, v = line.partition("=")
        v = v.strip().strip('"').strip("'")
        os.environ.setdefault(k.strip(), v)

# ── CONFIG ────────────────────────────────────────────────────────────────────

MODEL      = "claude-haiku-4-5-20251001"   # swap to claude-sonnet-4-6 for better accuracy
SCANS_DIR  = Path("scans")
OUTPUT_DIR = Path("output")
THUMB_DIR  = OUTPUT_DIR / "card_thumbs"
DB_PATH    = OUTPUT_DIR / "cards.db"
XLSX_PATH  = OUTPUT_DIR / "cards.xlsx"

THUMB_PX    = (280, 390)  # per-card thumbnail size — high-res for lightbox
THUMB_SM    = (110, 155)  # display size in Excel rows
EXCEL_ROW_H = 120         # Excel row height (points ≈ pixels * 0.75)
EXCEL_IMG_W = 17          # Excel image column width (chars)

# ── PROMPT ────────────────────────────────────────────────────────────────────

PROMPT = """\
Esta imagem contém duas cartas de Magic: The Gathering lado a lado.

Para CADA carta, identifique e retorne em português:
- nome        : nome da carta
- tipo        : tipo principal (Criatura, Feitiço, Mágica Instantânea, Encantamento, Artefato, Planeswalker, Terreno, etc.)
- subtipo     : subtipo se houver (ex: "Goblin Guerreiro"), ou null
- cor         : Branco, Azul, Preto, Vermelho, Verde, Incolor, ou Multicolorido
- custo_mana  : custo de mana em notação simbólica, ex: "{2}{W}{W}"  (W=Branco U=Azul B=Preto R=Vermelho G=Verde)
- texto       : texto de regras completo em português (traduza se estiver em inglês)
- poder_resist: para criaturas "X/Y", caso contrário null
- raridade    : Comum, Incomum, Rara, ou Mítica Rara
- edicao      : nome da edição/coleção se legível, caso contrário null

A carta da ESQUERDA é a primeira (índice 0), a da DIREITA é a segunda (índice 1).
Se uma carta estiver ilegível ou coberta, retorne null para todos os seus campos (exceto nome que pode ser "Ilegível").

Responda APENAS com JSON válido, sem texto adicional:
{
  "cards": [
    { "nome":"...", "tipo":"...", "subtipo":null, "cor":"...", "custo_mana":"...", "texto":"...", "poder_resist":null, "raridade":"...", "edicao":null },
    { "nome":"...", "tipo":"...", "subtipo":null, "cor":"...", "custo_mana":"...", "texto":"...", "poder_resist":null, "raridade":"...", "edicao":null }
  ]
}"""

# ── DATABASE ──────────────────────────────────────────────────────────────────

SCHEMA = """
CREATE TABLE IF NOT EXISTS cards (
    id            INTEGER PRIMARY KEY AUTOINCREMENT,
    imagem        TEXT,
    posicao       TEXT,
    thumb         TEXT,
    dono          TEXT,
    nome          TEXT,
    tipo          TEXT,
    subtipo       TEXT,
    cor           TEXT,
    custo_mana    TEXT,
    texto         TEXT,
    poder_resist  TEXT,
    raridade      TEXT,
    edicao        TEXT,
    criado_em     TEXT
);
CREATE TABLE IF NOT EXISTS imagens_processadas (
    nome TEXT PRIMARY KEY,
    em   TEXT
);
"""

def init_db(conn):
    for stmt in SCHEMA.strip().split(";"):
        s = stmt.strip()
        if s:
            conn.execute(s)
    # Migration: add dono column to existing databases
    cols = [r[1] for r in conn.execute("PRAGMA table_info(cards)").fetchall()]
    if "dono" not in cols:
        conn.execute("ALTER TABLE cards ADD COLUMN dono TEXT")
    conn.commit()

def is_done(conn, filename):
    return conn.execute(
        "SELECT 1 FROM imagens_processadas WHERE nome=?", (filename,)
    ).fetchone() is not None

def save_result(conn, filename, cards, thumbs, dono=None):
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    for card, pos, thumb in zip(cards, ["esquerda", "direita"], thumbs):
        conn.execute("""
            INSERT INTO cards
              (imagem, posicao, thumb, dono, nome, tipo, subtipo, cor, custo_mana,
               texto, poder_resist, raridade, edicao, criado_em)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            filename, pos, str(thumb) if thumb else None, dono,
            card.get("nome"), card.get("tipo"), card.get("subtipo"),
            card.get("cor"), card.get("custo_mana"), card.get("texto"),
            card.get("poder_resist"), card.get("raridade"), card.get("edicao"), ts
        ))
    conn.execute("INSERT INTO imagens_processadas VALUES (?,?)", (filename, ts))
    conn.commit()

# ── THUMBNAILS ────────────────────────────────────────────────────────────────

def make_thumbs(img_path, stem):
    """Split image left/right and save individual card thumbnails."""
    from PIL import ImageOps
    img = ImageOps.exif_transpose(PilImage.open(img_path))
    w, h = img.size
    mid = w // 2
    results = []
    for side, box in [("L", (0, 0, mid, h)), ("R", (mid, 0, w, h))]:
        half = img.crop(box)
        half.thumbnail(THUMB_PX, PilImage.LANCZOS)
        dest = THUMB_DIR / f"{stem}_{side}.jpg"
        half.save(dest, "JPEG", quality=82)
        results.append(dest)
    return results  # [left_path, right_path]

# ── CLAUDE API ────────────────────────────────────────────────────────────────

def identify(client, img_path):
    data = base64.standard_b64encode(img_path.read_bytes()).decode()
    resp = client.messages.create(
        model=MODEL,
        max_tokens=2048,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": "image/jpeg", "data": data}},
                {"type": "text", "text": PROMPT},
            ]
        }]
    )
    text = resp.content[0].text.strip()
    # Strip markdown fences if model wraps in ```json ... ```
    if text.startswith("```"):
        parts = text.split("```")
        text = parts[1].lstrip("json").strip() if len(parts) >= 2 else text
    return json.loads(text).get("cards", [{}, {}])

# ── EXCEL EXPORT ──────────────────────────────────────────────────────────────

HEADERS = ["#", "Imagem", "Dono", "Nome", "Tipo", "Subtipo", "Cor", "Custo Mana",
           "Texto / Habilidades", "P/R", "Raridade", "Edição", "Arquivo Origem"]
COL_WIDTHS = [5, EXCEL_IMG_W, 14, 26, 16, 20, 13, 13, 65, 7, 12, 22, 32]

def export_xlsx(conn):
    rows = conn.execute("""
        SELECT id, thumb, dono, nome, tipo, subtipo, cor, custo_mana,
               texto, poder_resist, raridade, edicao, imagem
        FROM cards ORDER BY dono, nome
    """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Coleção MTG"
    ws.freeze_panes = "A2"

    # Header
    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="95A5A6")
    border = Border(bottom=thin)

    for c, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(1, c, h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22

    alt_fill  = PatternFill("solid", fgColor="EBF5FB")
    norm_fill = PatternFill("solid", fgColor="FDFEFE")
    txt_font  = Font(size=10)
    wrap_align = Alignment(wrap_text=True, vertical="top")
    ctr_align  = Alignment(horizontal="center", vertical="top")

    for row_i, row in enumerate(rows, 2):
        id_, thumb, dono, nome, tipo, subtipo, cor, custo, texto, pr, raridade, edicao, src = row
        ws.row_dimensions[row_i].height = EXCEL_ROW_H

        fill = alt_fill if row_i % 2 == 0 else norm_fill
        vals = [id_, "", dono or "", nome, tipo, subtipo or "", cor or "", custo or "",
                texto or "", pr or "", raridade or "", edicao or "", src or ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row_i, c, v)
            cell.fill = fill
            cell.font = txt_font
            cell.alignment = ctr_align if c in (1, 6, 7, 9, 10) else wrap_align

        # Embed thumbnail (display at small size regardless of saved resolution)
        if thumb:
            tp = Path(thumb)
            if tp.is_file():
                try:
                    buf = BytesIO()
                    t = PilImage.open(tp)
                    t.thumbnail(THUMB_SM)
                    t.save(buf, "JPEG")
                    buf.seek(0)
                    xl_img = XLImage(buf)
                    xl_img.width  = THUMB_SM[0]
                    xl_img.height = THUMB_SM[1]
                    xl_img.anchor = f"B{row_i}"
                    ws.add_image(xl_img)
                except Exception as e:
                    ws.cell(row_i, 2, f"(err: {e})")

    wb.save(XLSX_PATH)
    print(f"  Excel: {XLSX_PATH}  ({len(rows)} cartas)")

# ── HTML GALLERY ──────────────────────────────────────────────────────────────

def export_html(conn):
    rows = conn.execute("""
        SELECT id, thumb, nome, tipo, subtipo, cor, custo_mana,
               raridade, edicao, poder_resist, texto, dono
        FROM cards ORDER BY nome
    """).fetchall()

    cards_json = json.dumps([{
        "id":      r[0],
        "thumb":   Path(r[1]).name if r[1] else "",
        "nome":    r[2]  or "",
        "tipo":    r[3]  or "",
        "subtipo": r[4]  or "",
        "cor":     r[5]  or "",
        "custo":   r[6]  or "",
        "raridade":r[7]  or "",
        "edicao":  r[8]  or "",
        "pr":      r[9]  or "",
        "texto":   r[10] or "",
        "dono":    r[11] or "",
    } for r in rows], ensure_ascii=False)

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>Coleção MTG</title>
<style>
  :root{{--bg:#1a1a2e;--panel:#16213e;--accent:#e94560;--text:#eaeaea;--dim:#888;--border:#0f3460}}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:var(--bg);color:var(--text);font-family:'Segoe UI',system-ui,sans-serif;padding:1.5rem}}
  h1{{font-size:1.8rem;margin-bottom:.5rem}}h1 span{{color:var(--accent)}}
  .stats{{color:var(--dim);font-size:.85rem;margin-bottom:1.2rem}}
  #search{{width:100%;max-width:400px;padding:.5rem .8rem;border-radius:8px;border:1px solid var(--border);
           background:var(--panel);color:var(--text);font-size:1rem;margin-bottom:1.2rem}}
  .grid{{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:1rem}}
  .card{{background:var(--panel);border:1px solid var(--border);border-radius:10px;overflow:hidden;
          cursor:pointer;transition:transform .15s;user-select:none}}
  .card:hover{{transform:scale(1.03);border-color:var(--accent)}}
  .card img{{width:100%;display:block;aspect-ratio:7/10;object-fit:cover;background:#111}}
  .card-info{{padding:.55rem .6rem}}
  .card-name{{font-size:.82rem;font-weight:600;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
  .card-meta{{font-size:.72rem;color:var(--dim);margin-top:.2rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
  .hint{{font-size:.65rem;color:#555;margin-top:.15rem}}
  /* owner filters */
  .filters{{display:flex;flex-wrap:wrap;gap:.5rem;margin-bottom:1.2rem}}
  .filter-btn{{padding:.35rem .9rem;border-radius:20px;border:1px solid var(--border);
               background:var(--panel);color:var(--dim);cursor:pointer;font-size:.82rem;
               transition:background .15s,color .15s}}
  .filter-btn:hover{{border-color:var(--accent);color:var(--text)}}
  .filter-btn.active{{background:var(--accent);border-color:var(--accent);color:#fff;font-weight:600}}
  /* colour badge */
  .badge{{display:inline-block;padding:.1rem .4rem;border-radius:4px;font-size:.68rem;margin-right:.3rem}}
  .Branco{{background:#f0e68c;color:#333}}.Azul{{background:#4169e1;color:#fff}}
  .Preto{{background:#2f2f2f;color:#ccc;border:1px solid #555}}.Vermelho{{background:#c0392b;color:#fff}}
  .Verde{{background:#27ae60;color:#fff}}.Incolor{{background:#aaa;color:#333}}
  .Multicolorido{{background:linear-gradient(90deg,#f1c40f,#27ae60);color:#333}}
  /* detail modal */
  .modal{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.7);z-index:10;
          align-items:center;justify-content:center}}
  .modal.open{{display:flex}}
  .modal-box{{background:var(--panel);border:1px solid var(--border);border-radius:14px;
               padding:1.5rem;max-width:520px;width:90%;position:relative;max-height:90vh;overflow-y:auto}}
  .modal-box h2{{font-size:1.2rem;color:var(--accent);margin-bottom:.8rem}}
  .modal-thumb{{float:right;margin:0 0 .8rem 1rem;border-radius:6px;width:100px;cursor:zoom-in}}
  .modal-box table{{width:100%;border-collapse:collapse;font-size:.85rem;clear:both}}
  .modal-box td{{padding:.35rem .4rem;vertical-align:top;border-bottom:1px solid var(--border)}}
  .modal-box td:first-child{{color:var(--dim);width:36%;font-size:.78rem}}
  .texto-cell{{font-size:.8rem;line-height:1.5;white-space:pre-wrap}}
  .close{{position:absolute;top:.7rem;right:.9rem;font-size:1.4rem;cursor:pointer;color:var(--dim)}}
  .close:hover{{color:var(--accent)}}
  /* lightbox */
  #lightbox{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.92);z-index:20;
             align-items:center;justify-content:center;cursor:zoom-out}}
  #lightbox.open{{display:flex}}
  #lb-box{{text-align:center}}
  #lb-name{{color:#eee;font-size:1.1rem;font-weight:600;margin-bottom:.6rem;
            text-shadow:0 1px 4px #000}}
  #lb-img{{max-height:88vh;max-width:92vw;border-radius:10px;
           box-shadow:0 8px 48px rgba(0,0,0,.9);
           user-select:none;-webkit-user-select:none;pointer-events:none}}
</style>
</head>
<body>
<h1>Coleção <span>MTG</span></h1>
<p class="stats" id="count"></p>
<input id="search" type="text" placeholder="Buscar por nome, tipo, cor, raridade...">
<div class="filters" id="filters"></div>
<div class="grid" id="grid"></div>

<!-- detail modal -->
<div class="modal" id="modal">
  <div class="modal-box" id="modal-box">
    <span class="close" id="close">&times;</span>
    <h2 id="m-nome"></h2>
    <img class="modal-thumb" id="m-img" src="" alt="" title="Duplo clique para ampliar">
    <table id="m-table"></table>
  </div>
</div>

<!-- full-size lightbox -->
<div id="lightbox">
  <div id="lb-box">
    <div id="lb-name"></div>
    <img id="lb-img" src="" alt="">
  </div>
</div>

<script>
const DATA = {cards_json};
const thumbDir = "card_thumbs/";

// ── Mana symbol renderer ──────────────────────────────────────────────────────
const MANA_BG = {{W:'#f9f6d5',U:'#3a6bc4',B:'#1a1a1a',R:'#c0392b',G:'#1e8449',C:'#bdc3c7',X:'#7f8c8d',S:'#a8d8ea'}};
const MANA_FG = {{W:'#7a6a1b',U:'#fff',B:'#bbb',R:'#fff',G:'#fff',C:'#444',X:'#fff',S:'#333'}};

function manaHtml(cost) {{
  if (!cost) return '—';
  return (cost.match(/\\{{([^}}]+)\\}}/g) || []).map(m => {{
    const s = m.slice(1,-1);
    const bg = MANA_BG[s] || '#7f8c8d';
    const fg = MANA_FG[s] || '#fff';
    return `<span style="display:inline-flex;align-items:center;justify-content:center;`+
           `width:22px;height:22px;border-radius:50%;background:${{bg}};color:${{fg}};`+
           `font-size:11px;font-weight:800;margin:1px;border:1.5px solid rgba(0,0,0,.35);`+
           `font-family:monospace;flex-shrink:0">${{s}}</span>`;
  }}).join('');
}}

// ── Owner filter ──────────────────────────────────────────────────────────────
let activeDono = "Todos";

function buildFilters() {{
  const owners = ["Todos", ...new Set(DATA.map(c => c.dono).filter(Boolean).sort())];
  const container = document.getElementById("filters");
  owners.forEach(o => {{
    const btn = document.createElement("button");
    btn.className = "filter-btn" + (o === "Todos" ? " active" : "");
    btn.textContent = o === "Todos" ? `Todos (${{DATA.length}})` :
      `${{o}} (${{DATA.filter(c => c.dono === o).length}})`;
    btn.onclick = () => {{
      activeDono = o;
      document.querySelectorAll(".filter-btn").forEach(b => b.classList.remove("active"));
      btn.classList.add("active");
      applyFilters();
    }};
    container.appendChild(btn);
  }});
}}

function applyFilters() {{
  const q = document.getElementById("search").value.toLowerCase();
  let list = activeDono === "Todos" ? DATA : DATA.filter(c => c.dono === activeDono);
  if (q) list = list.filter(c =>
    (c.nome + c.tipo + c.cor + c.edicao + c.raridade + c.subtipo + c.dono)
      .toLowerCase().includes(q));
  render(list);
}}

// ── Grid render ───────────────────────────────────────────────────────────────
function render(list) {{
  document.getElementById("count").textContent = list.length + " carta(s) — clique para detalhes · duplo clique para ampliar";
  const grid = document.getElementById("grid");
  grid.innerHTML = "";
  list.forEach(c => {{
    const div = document.createElement("div");
    div.className = "card";
    div.innerHTML =
      `<img src="${{thumbDir}}${{c.thumb}}" alt="${{c.nome}}" loading="lazy" onerror="this.style.display='none'">
       <div class="card-info">
         <div class="card-name">${{c.nome}}</div>
         <div class="card-meta"><span class="badge ${{c.cor}}">${{c.cor}}</span>${{c.tipo}}</div>
         <div class="card-meta">${{c.edicao || ''}}</div>
       </div>`;
    div.onclick    = ()  => openModal(c);
    div.ondblclick = (e) => {{ e.stopPropagation(); openLightbox(c); }};
    grid.appendChild(div);
  }});
}}

// ── Detail modal ──────────────────────────────────────────────────────────────
function openModal(c) {{
  document.getElementById("m-nome").textContent = c.nome;
  const img = document.getElementById("m-img");
  img.src = thumbDir + c.thumb;
  img.ondblclick = (e) => {{ e.stopPropagation(); openLightbox(c); }};
  const tableRows = [
    ["Dono",       `<strong>${{c.dono || "—"}}</strong>`],
    ["Tipo",       c.tipo],
    ["Subtipo",    c.subtipo || "—"],
    ["Cor",        c.cor],
    ["Custo Mana", manaHtml(c.custo)],
    ["P/R",        c.pr || "—"],
    ["Raridade",   c.raridade],
    ["Edição",     c.edicao || "—"],
    ["Habilidades",`<span class='texto-cell'>${{c.texto || "—"}}</span>`],
  ];
  document.getElementById("m-table").innerHTML =
    tableRows.map(([k,v]) => `<tr><td>${{k}}</td><td>${{v || "—"}}</td></tr>`).join("");
  document.getElementById("modal").classList.add("open");
}}

document.getElementById("close").onclick = () =>
  document.getElementById("modal").classList.remove("open");
document.getElementById("modal").onclick = e => {{
  if (!document.getElementById("modal-box").contains(e.target))
    document.getElementById("modal").classList.remove("open");
}};

// ── Lightbox ──────────────────────────────────────────────────────────────────
function openLightbox(c) {{
  document.getElementById("lb-img").src  = thumbDir + c.thumb;
  document.getElementById("lb-name").textContent = c.nome;
  document.getElementById("lightbox").classList.add("open");
}}

document.getElementById("lightbox").onclick = () =>
  document.getElementById("lightbox").classList.remove("open");

document.addEventListener("keydown", e => {{
  if (e.key === "Escape") {{
    document.getElementById("modal").classList.remove("open");
    document.getElementById("lightbox").classList.remove("open");
  }}
}});

// ── Search ────────────────────────────────────────────────────────────────────
document.getElementById("search").addEventListener("input", applyFilters);

buildFilters();
applyFilters();
</script>
</body>
</html>"""

    html_path = OUTPUT_DIR / "colecao.html"
    html_path.write_text(html, encoding="utf-8")
    print(f"  HTML:  {html_path}  ({len(rows)} cartas)")

# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    # Force UTF-8 output on Windows so card names with accents/symbols print correctly
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    load_dotenv()

    OUTPUT_DIR.mkdir(exist_ok=True)
    THUMB_DIR.mkdir(exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    init_db(conn)

    # --rethumb: regenerate all thumbnails at new resolution, then re-export (no API key needed)
    if "--rethumb" in sys.argv:
        done = conn.execute("SELECT nome FROM imagens_processadas").fetchall()
        console.print(f"[bold]Regenerando thumbnails[/bold] {THUMB_PX[0]}×{THUMB_PX[1]}px — {len(done)} imagens")
        with Progress(SpinnerColumn(), BarColumn(bar_width=40), MofNCompleteColumn(),
                      TimeElapsedColumn(), console=console) as prog:
            task = prog.add_task("thumbnails", total=len(done))
            for (fname,) in done:
                img_path = SCANS_DIR / fname
                if img_path.is_file():
                    thumbs = make_thumbs(img_path, img_path.stem)
                    conn.execute("UPDATE cards SET thumb=? WHERE imagem=? AND posicao='esquerda'",
                                 (str(thumbs[0]), fname))
                    conn.execute("UPDATE cards SET thumb=? WHERE imagem=? AND posicao='direita'",
                                 (str(thumbs[1]), fname))
                prog.advance(task)
        conn.commit()
        console.print("[green]Pronto.[/green] Exportando...")
        export_xlsx(conn)
        export_html(conn)
        conn.close()
        return

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERRO: defina a variável ANTHROPIC_API_KEY antes de executar.")
        print('  PowerShell: $env:ANTHROPIC_API_KEY = "sk-ant-..."')
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    images = sorted(SCANS_DIR.glob("*.jpg")) + sorted(SCANS_DIR.glob("*.jpeg"))
    if not images:
        print(f"Nenhuma imagem encontrada em {SCANS_DIR.resolve()}")
        sys.exit(0)

    # Resolve owner — --dono "Name" flag or interactive prompt
    dono_arg = next((sys.argv[i+1] for i, a in enumerate(sys.argv)
                     if a == "--dono" and i+1 < len(sys.argv)), None)

    todo = [p for p in images if not is_done(conn, p.name)]
    skipped = len(images) - len(todo)

    if todo and not dono_arg:
        console.print("[yellow]Dono das cartas não especificado.[/yellow]")
        console.print("  Use [bold]--dono \"Nome\"[/bold] ou digite abaixo.")
        dono_arg = console.input("  Dono: ").strip() or "Desconhecido"

    console.print(f"[bold]MTG Card Identifier[/bold] — modelo: [cyan]{MODEL}[/cyan]")
    if todo:
        console.print(f"  Dono: [bold magenta]{dono_arg}[/bold magenta]")
    console.print(f"  {len(images)} imagens  |  [green]{len(todo)} a processar[/green]  |  "
                  f"[dim]{skipped} já processadas[/dim]\n")

    gui_mode = "--gui" in sys.argv

    if todo:
        if gui_mode:
            for i, img_path in enumerate(todo, 1):
                thumbs = make_thumbs(img_path, img_path.stem)
                try:
                    cards = identify(client, img_path)
                    while len(cards) < 2:
                        cards.append({})
                    save_result(conn, img_path.name, cards[:2], thumbs, dono=dono_arg)
                    names = [c.get("nome", "?") or "?" for c in cards[:2]]
                    print(f"PROG:{i}/{len(todo)}:{names[0]}|{names[1]}", flush=True)
                except json.JSONDecodeError:
                    print(f"PROG:{i}/{len(todo)}:ERRO JSON|", flush=True)
                except anthropic.APIError as e:
                    print(f"PROG:{i}/{len(todo)}:ERRO API {e.status_code}|", flush=True)
                except Exception as e:
                    print(f"PROG:{i}/{len(todo)}:ERRO|{e}", flush=True)
                time.sleep(0.4)
        else:
            with Progress(
                SpinnerColumn(),
                TextColumn("[progress.description]{task.description}"),
                BarColumn(bar_width=38),
                MofNCompleteColumn(),
                TimeElapsedColumn(),
                TextColumn("ETA"),
                TimeRemainingColumn(),
                TextColumn("[dim]{task.fields[last]}", justify="right"),
                console=console,
            ) as progress:
                task = progress.add_task("[cyan]Identificando", total=len(todo), last="")

                for img_path in todo:
                    thumbs = make_thumbs(img_path, img_path.stem)
                    try:
                        cards = identify(client, img_path)
                        while len(cards) < 2:
                            cards.append({})
                        save_result(conn, img_path.name, cards[:2], thumbs, dono=dono_arg)
                        names = [c.get("nome", "?") or "?" for c in cards[:2]]
                        last = f"{names[0]}  |  {names[1]}"
                    except json.JSONDecodeError:
                        last = "[red]JSON error[/red]"
                    except anthropic.APIError as e:
                        last = f"[red]API error: {e.status_code}[/red]"
                    except Exception as e:
                        last = f"[red]{e}[/red]"

                    progress.update(task, advance=1, last=last)
                    time.sleep(0.4)

    total = conn.execute("SELECT COUNT(*) FROM cards").fetchone()[0]
    console.print(f"\n[bold green]Concluído[/bold green] — {total} cartas no banco\n")
    console.print("Exportando...")
    export_xlsx(conn)
    export_html(conn)
    conn.close()
    console.print("\n[bold green]Pronto.[/bold green]")

if __name__ == "__main__":
    main()
